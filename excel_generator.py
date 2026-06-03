#!/usr/bin/env python3
"""
Excel generator — fills Javier's pricing_and_calc.xlsx template with
extraction data. Derived quantities and Computo totals are written as
plain values so the sheet displays correctly without any formula recalculation.
"""

import os
import re
import shutil
import tempfile
from pathlib import Path
from typing import Any, Dict

TEMPLATE_PATH = (
    Path(__file__).parent / "javier-files" / "pricing_and_calc" / "pricing_and_calc.xlsx"
)

# Unit cost prices (EUR) — must match PriceList sheet
_PRICES = {
    "STR-001": 285,   # X-lam ext wall 140mm / m2
    "STR-002": 230,   # X-lam int wall 100mm / m2
    "STR-010": 165,   # GL24h beam 200x400 / m
    "STR-011": 245,   # GL24h beam 240x480 / m
    "STR-020": 315,   # X-lam intermediate slab / m2
    "STR-021": 295,   # X-lam roof slab / m2
    "ENV-001":  78,   # Wall insulation 200mm / m2
    "ENV-002":  92,   # Roof insulation 240mm / m2
    "ENV-003": 145,   # Ventilated facade larch / m2
    "ENV-004":  95,   # Ceramic tile roof package / m2
    "ENV-005":  48,   # Flashings + gutter / m
    "WIN-001": 720,   # Triple-glazing window / m2
    "WIN-010": 380,   # External shutter / unit
    "WIN-020": 1850,  # Glazed entrance door / unit
    "SITE-01": 2200,  # Crane rental / month
    "SITE-02":  950,  # Scaffolding / month
    "SITE-10": 18000, # Engineering lump sum
    "SITE-20":  780,  # Assembly labour / day
}

_COMPUTO_ORDER = [
    "STR-001", "STR-002", "STR-010", "STR-011", "STR-020", "STR-021",
    "ENV-001", "ENV-002", "ENV-003", "ENV-004", "ENV-005",
    "WIN-001", "WIN-010", "WIN-020",
    "SITE-01", "SITE-02", "SITE-10", "SITE-20",
]

_CATEGORIES = {
    "STR-001": "STR", "STR-002": "STR", "STR-010": "STR", "STR-011": "STR",
    "STR-020": "STR", "STR-021": "STR",
    "ENV-001": "ENV", "ENV-002": "ENV", "ENV-003": "ENV", "ENV-004": "ENV",
    "ENV-005": "ENV",
    "WIN-001": "WIN", "WIN-010": "WIN", "WIN-020": "WIN",
    "SITE-01": "SITE", "SITE-02": "SITE", "SITE-10": "SITE", "SITE-20": "SITE",
}

_DESCRIPTIONS = {
    "STR-001": "X-lam wall panel 140 mm (external load-bearing)",
    "STR-002": "X-lam wall panel 100 mm (internal load-bearing)",
    "STR-010": "GL24h glulam beam 200x400",
    "STR-011": "GL24h glulam beam 240x480 (large spans)",
    "STR-020": "X-lam intermediate slab 200 mm",
    "STR-021": "X-lam roof slab 160 mm",
    "ENV-001": "Wood-fibre insulation 200 mm (walls)",
    "ENV-002": "Wood-fibre insulation 240 mm (roof)",
    "ENV-003": "Ventilated facade - larch cladding",
    "ENV-004": "Ceramic tile roof package (membrane + tiles)",
    "ENV-005": "Aluminium flashings and gutter set",
    "WIN-001": "Aluminium-clad timber window, triple glazing",
    "WIN-010": "External louvre shutter (manual)",
    "WIN-020": "Glazed aluminium entrance door",
    "SITE-01": "Site crane rental (incl. operator, per month)",
    "SITE-02": "Facade scaffolding (per month)",
    "SITE-10": "Engineering and project management (lump sum)",
    "SITE-20": "On-site assembly labour",
}

_ASSEMBLY_DAYS = 35
_SITE_MONTHS   = 3


def _num(value, default=0):
    try:
        return float(value) if value is not None else default
    except (TypeError, ValueError):
        return default


def generate_excel(extraction_data: Dict[str, Any]) -> bytes:
    """
    Populate the pricing_and_calc.xlsx template from extraction JSON.

    Writes input cells to Parameters and Schedule, pre-computes all derived
    quantities (ext/int wall area, slab, beams) and writes them as plain values
    to Schedule rows 20-25, then writes all Computo line-item quantities,
    unit prices, and totals as values. No formula recalculation required.

    Args:
        extraction_data: full extraction dict (may be wrapped under 'extraction' key)

    Returns:
        Raw .xlsx bytes ready to send as a download.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    ext = extraction_data.get("extraction") or extraction_data

    dims         = ext.get("building_dimensions") or {}
    geo          = ext.get("building_geometry") or {}
    specs        = ext.get("technical_specifications") or {}
    commercial   = ext.get("commercial_terms") or {}
    struct_notes = ext.get("structural_notes") or []

    g0    = geo.get("ground_floor") or {}
    g1    = geo.get("first_floor") or {}
    attic = geo.get("attic") or {}
    roof  = geo.get("roof") or {}
    wins  = geo.get("windows") or {}

    wall_sys  = specs.get("wall_system") or {}
    ins_walls = (specs.get("insulation") or {}).get("walls") or {}
    ins_roof  = (specs.get("insulation") or {}).get("roof") or {}

    # ── Pre-compute derived geometry ─────────────────────────────────
    floor_height  = 2.8
    # Prefer explicit min height for conservative costing; fall back to midpoint, then default 1.5m
    attic_height  = (
        _num(attic.get("height_min_m"), None) or
        _num(attic.get("height_m"), 1.5)
    )
    perims        = [_num(g0.get("ext_perimeter_m")),
                     _num(g1.get("ext_perimeter_m")),
                     _num(attic.get("ext_perimeter_m"), 0)]
    int_walls_len = [_num(g0.get("int_walls_length_m")),
                     _num(g1.get("int_walls_length_m")),
                     0]
    heights       = [floor_height, floor_height, attic_height]

    ext_wall_m2      = sum(p * h for p, h in zip(perims, heights))
    int_wall_m2      = sum(l * h for l, h in zip(int_walls_len, heights))
    p0_hab           = _num(g0.get("habitable_area_m2") or dims.get("ground_floor_area_m2"))
    p1_hab           = _num(g1.get("habitable_area_m2") or dims.get("first_floor_area_m2"))
    attic_m2         = _num(attic.get("area_m2") or dims.get("attic_area_m2"))
    intermediate_slab = p1_hab + attic_m2
    roof_m2          = _num(roof.get("projected_area_m2"))
    gutter_m         = _num(roof.get("gutter_length_m"))
    window_m2        = _num(wins.get("total_area_m2"))
    shutter_count    = _num(wins.get("count"), 0)

    # Large-span beam — only cost when engineering decision is resolved/confirmed/approved.
    # Unresolved (flagged for decision / pending / unresolved) → 0 so the line item is
    # excluded from totals rather than silently priced at a wrong dimension.
    _RESOLVED = {"approved", "resolved", "confirmed"}
    large_span_m = 0.0
    for note in struct_notes:
        if (note.get("status") or "").lower().strip() not in _RESOLVED:
            continue
        for field in ("constraint", "proposed_solution"):
            m = re.search(r'(\d+(?:\.\d+)?)\s*m\b', note.get(field, ""))
            if m:
                large_span_m = float(m.group(1))
                break
        else:
            continue
        break

    std_beam_m = intermediate_slab * 0.18

    # Quantities per line item (same order as Computo rows 5-22)
    quantities = {
        "STR-001": ext_wall_m2,
        "STR-002": int_wall_m2,
        "STR-010": std_beam_m,
        "STR-011": large_span_m,
        "STR-020": intermediate_slab,
        "STR-021": roof_m2,
        "ENV-001": ext_wall_m2,
        "ENV-002": roof_m2,
        "ENV-003": ext_wall_m2,
        "ENV-004": roof_m2,
        "ENV-005": gutter_m,
        "WIN-001": window_m2,
        "WIN-010": shutter_count,
        "WIN-020": 1,
        "SITE-01": _SITE_MONTHS,
        "SITE-02": _SITE_MONTHS,
        "SITE-10": 1,
        "SITE-20": _ASSEMBLY_DAYS,
    }

    margin  = _num(commercial.get("margin_percent"), 30) / 100
    vat     = _num(commercial.get("vat_rate_percent"), 21) / 100

    # ── Aggregation totals ───────────────────────────────────────────
    totals_by_cat = {cat: 0.0 for cat in ("STR", "ENV", "WIN", "SITE")}
    for code in _COMPUTO_ORDER:
        totals_by_cat[_CATEGORIES[code]] += quantities[code] * _PRICES[code]

    total_cost    = sum(totals_by_cat.values())
    selling_price = total_cost / (1 - margin) if margin < 1 else total_cost
    vat_amount    = selling_price * vat
    total_incl_vat = selling_price + vat_amount

    # ── Write to workbook ────────────────────────────────────────────
    tmp_dir = tempfile.mkdtemp()
    try:
        tmp_path = os.path.join(tmp_dir, "pricing_and_calc.xlsx")
        shutil.copy2(TEMPLATE_PATH, tmp_path)
        wb = openpyxl.load_workbook(tmp_path)

        # ── Parameters sheet ─────────────────────────────────────────
        params = wb["Parameters"]
        params["B4"]  = round(margin, 4)
        params["B5"]  = round(vat, 4)
        params["B7"]  = floor_height
        params["B8"]  = _num(wall_sys.get("thickness_mm"), 140)
        params["B10"] = _num(ins_walls.get("thickness_mm"), 200)
        params["B11"] = _num(ins_roof.get("thickness_mm"), 240)
        params["B12"] = 0.3   # terrace coefficient
        params["B13"] = 0.0   # roof coefficient
        params["B14"] = _SITE_MONTHS

        # ── Schedule sheet — input cells ─────────────────────────────
        sched = wb["Schedule"]

        sched["B5"]  = p0_hab
        sched["C5"]  = _num(g0.get("terrace_area_m2"))
        sched["D5"]  = _num(g0.get("ext_perimeter_m"))
        sched["E5"]  = _num(g0.get("int_walls_length_m"))

        sched["B6"]  = p1_hab
        sched["C6"]  = _num(g1.get("terrace_area_m2") or dims.get("terrace_area_m2"))
        sched["D6"]  = _num(g1.get("ext_perimeter_m"))
        sched["E6"]  = _num(g1.get("int_walls_length_m"))

        sched["B7"]  = attic_m2
        sched["D7"]  = _num(attic.get("ext_perimeter_m"), 0)
        sched["F7"]  = attic_height

        sched["B11"] = roof_m2
        sched["B12"] = gutter_m
        sched["B15"] = window_m2
        sched["B16"] = shutter_count
        sched["B17"] = 1

        # Derived quantities — write as values so Computo sees numbers immediately
        sched["B20"] = round(ext_wall_m2, 2)
        sched["B21"] = round(int_wall_m2, 2)
        sched["B22"] = round(intermediate_slab, 2)
        sched["B24"] = round(std_beam_m, 2)
        sched["B25"] = large_span_m

        # ── Computo sheet — write all quantities, prices, and totals ──
        comp = wb["Computo"]

        for i, code in enumerate(_COMPUTO_ORDER):
            row = 5 + i
            qty   = round(quantities[code], 2)
            price = _PRICES[code]
            comp[f"B{row}"] = _DESCRIPTIONS[code]
            comp[f"D{row}"] = qty
            comp[f"F{row}"] = price
            comp[f"G{row}"] = round(qty * price, 2)

        # Aggregation section
        comp["G27"] = round(totals_by_cat["STR"], 2)
        comp["G28"] = round(totals_by_cat["ENV"], 2)
        comp["G29"] = round(totals_by_cat["WIN"], 2)
        comp["G30"] = round(totals_by_cat["SITE"], 2)
        comp["G31"] = round(total_cost, 2)
        comp["G32"] = round(margin, 4)
        comp["G33"] = round(selling_price, 2)
        comp["G34"] = round(vat_amount, 2)
        comp["G35"] = round(total_incl_vat, 2)

        # ── Audit sheet — human-readable formula traces ──────────────
        audit = wb.create_sheet("Audit")
        audit["A1"] = "Line item"
        audit["B1"] = "Quantity"
        audit["C1"] = "Unit"
        audit["D1"] = "Formula / source"
        audit_rows = [
            ("STR-001", round(quantities["STR-001"], 2), "m²",
             f"ext_wall = ({round(perims[0],1)}+{round(perims[1],1)}+{round(perims[2],1)}) perims × ({floor_height},{floor_height},{round(attic_height,2)}) heights"),
            ("STR-002", round(quantities["STR-002"], 2), "m²",
             f"int_wall = ({round(int_walls_len[0],1)}+{round(int_walls_len[1],1)}) × {floor_height}m"),
            ("STR-010", round(quantities["STR-010"], 2), "m",
             f"std_beam = intermediate_slab({round(intermediate_slab,2)}) × 0.18"),
            ("STR-011", round(quantities["STR-011"], 2), "m",
             f"large_span = {large_span_m}m from structural_notes (0 = unresolved/excluded)"),
            ("STR-020", round(quantities["STR-020"], 2), "m²",
             f"intermediate_slab = p1_hab({round(p1_hab,1)}) + attic({round(attic_m2,1)})"),
            ("STR-021", round(quantities["STR-021"], 2), "m²",
             f"roof projected area"),
            ("ENV-001", round(quantities["ENV-001"], 2), "m²", "= ext_wall_m2"),
            ("ENV-002", round(quantities["ENV-002"], 2), "m²", "= roof_m2"),
            ("ENV-003", round(quantities["ENV-003"], 2), "m²", "= ext_wall_m2"),
            ("ENV-004", round(quantities["ENV-004"], 2), "m²", "= roof_m2"),
            ("ENV-005", round(quantities["ENV-005"], 2), "m",  "= gutter_length_m"),
            ("WIN-001", round(quantities["WIN-001"], 2), "m²", "= windows.total_area_m2"),
            ("WIN-010", round(quantities["WIN-010"], 2), "u",  "= windows.count"),
            ("WIN-020", 1,                               "u",  "fixed = 1 entrance door"),
            ("SITE-01", _SITE_MONTHS,                    "mo", f"fixed = {_SITE_MONTHS} months"),
            ("SITE-02", _SITE_MONTHS,                    "mo", f"fixed = {_SITE_MONTHS} months"),
            ("SITE-10", 1,                               "ls", "fixed = engineering lump sum"),
            ("SITE-20", _ASSEMBLY_DAYS,                  "d",  f"fixed = {_ASSEMBLY_DAYS} assembly days"),
        ]
        for i, row_data in enumerate(audit_rows, start=2):
            for j, val in enumerate(row_data, start=1):
                audit.cell(row=i, column=j, value=val)
        # Summary
        audit[f"A{len(audit_rows)+3}"] = "total_cost"
        audit[f"B{len(audit_rows)+3}"] = round(total_cost, 2)
        audit[f"A{len(audit_rows)+4}"] = "margin %"
        audit[f"B{len(audit_rows)+4}"] = round(margin * 100, 1)
        audit[f"A{len(audit_rows)+5}"] = "selling_price"
        audit[f"B{len(audit_rows)+5}"] = round(selling_price, 2)
        audit[f"A{len(audit_rows)+6}"] = "vat_amount"
        audit[f"B{len(audit_rows)+6}"] = round(vat_amount, 2)
        audit[f"A{len(audit_rows)+7}"] = "total_incl_vat"
        audit[f"B{len(audit_rows)+7}"] = round(total_incl_vat, 2)

        wb.calculation.fullCalcOnLoad = True
        wb.save(tmp_path)
        return Path(tmp_path).read_bytes()

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
